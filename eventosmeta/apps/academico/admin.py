
"""
Admin do app ACADÊMICO
Arquivo: apps/academico/admin.py
Finalidade: Configuração do admin para models de acadêmico
Atualizações:
    - 11/12/2025 - Adicionado seletor de cor visual e removido código hex da listagem
    - 20/01/2026 - Registrados todos os models no admin.site customizado (melhor prática)
    - 03/02/2026 - Adicionado filtro por evento e action para certificados
                - Sistema completo de certificados com 2 opções
                - Solução híbrida com filtros superiores + edição em massa
                - Adicionadas actions para relatório Excel e PDF
                - Título junto com logo, checkbox corrigido, labels removidos
    - 10/06/2026 - Refatoração para usar services
    - 13/07/2026 - CORRIGIDO: Removido import de admin_site customizado
                - Usando admin.site padrão do Django
                - Adicionado comentário explicativo
                - Corrigido erro de importação                
"""

from django import forms
from django.contrib import admin
from django.utils.html import format_html
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib import messages
from django.urls import reverse
from datetime import date
from collections import defaultdict

from .models import StatusMatricula, Matricula, Avaliacao
from apps.admin_mixins import CustomTitleMixin


# ==========================================
# STATUS MATRÍCULA
# ==========================================

class StatusMatriculaForm(forms.ModelForm):
    """Form personalizado com seletor de cor"""
    class Meta:
        model = StatusMatricula
        fields = '__all__'
        widgets = {
            'cor': forms.TextInput(attrs={
                'type': 'color',
                'style': 'width: 100px; height: 40px; cursor: pointer; border: 2px solid #ccc; border-radius: 4px;'
            })
        }


@admin.register(StatusMatricula)
class StatusMatriculaAdmin(CustomTitleMixin, admin.ModelAdmin):
    custom_title = 'adm Status de Matrícula'
    form = StatusMatriculaForm
    list_display = ['nome', 'cor_display', 'ordem']
    search_fields = ['nome']
    ordering = ['ordem', 'nome']

    fieldsets = (
        (None, {
            'fields': ('nome', 'cor', 'ordem'),
        }),
    )

    def cor_display(self, obj):
        """Exibe quadrado colorido"""
        if obj.cor:
            return format_html(
                '<span style="display: inline-block; width: 30px; height: 30px; '
                'background-color: {}; border: 2px solid #ccc; border-radius: 4px;"></span>',
                obj.cor
            )
        return '—'
    cor_display.short_description = 'Cor'
    cor_display.admin_order_field = 'cor'


# ==========================================
# MATRÍCULA
# ==========================================

@admin.register(Matricula)
class MatriculaAdmin(CustomTitleMixin, admin.ModelAdmin):
    custom_title = 'adm Matriculas'

    list_display = [
        'numero_matricula',
        'get_interessado',
        # 'get_evento',
        'turma',
        'status',
        'data_matricula'
    ]
    list_filter = [
        'status',
        'turma__evento',
        'turma',
        'data_matricula'
    ]
    search_fields = [
        'numero_matricula',
        'interessado__nome',
        'interessado__cpf',
        'turma__nome',
        'turma__evento__nome'
    ]
    date_hierarchy = 'data_matricula'
    ordering = ['-data_matricula']

    fieldsets = (
        ('Matrícula', {
            'fields': ('numero_matricula', 'turma', 'status')
        }),
        ('Dados do Aluno', {
            'fields': ('interessado', 'inscricao')
        }),
        ('Observações', {
            'fields': ('observacoes',)
        }),
        ('Auditoria', {
            'fields': ('data_matricula', 'data_atualizacao'),
            'classes': ('collapse',)
        }),
    )

    readonly_fields = ['numero_matricula', 'data_matricula', 'data_atualizacao']

    def get_interessado(self, obj):
        return obj.interessado.nome
    get_interessado.short_description = 'Interessado'
    get_interessado.admin_order_field = 'interessado__nome'

    def get_evento(self, obj):
        return obj.turma.evento.nome
    get_evento.short_description = 'Evento'
    get_evento.admin_order_field = 'turma__evento__nome'


# ==========================================
# AVALIAÇÃO
# ==========================================

@admin.register(Avaliacao)
class AvaliacaoAdmin(CustomTitleMixin, admin.ModelAdmin):
    custom_title = 'adm Notas e Frequência'
    change_list_template = 'admin/academico/avaliacao/digitar_notas.html'
    
    list_display = [
        'get_numero_matricula',
        'get_aluno',
        'get_turma',
        'nota_final',
        'frequencia',
        'aprovado',
        'certificado_emitido',
        'acoes_certificado'
    ]
    
    list_editable = [
        'nota_final',
        'frequencia',
        'aprovado'
    ]
    
    list_filter = [
        ('matricula__turma__evento', admin.RelatedOnlyFieldListFilter),
        ('matricula__turma', admin.RelatedOnlyFieldListFilter),
        'aprovado',
        'certificado_emitido',
    ]
    
    search_fields = [
        'matricula__numero_matricula',
        'matricula__interessado__nome',
        'matricula__interessado__cpf'
    ]
    
    ordering = ['matricula__interessado__nome']
    preserve_filters = True
    
    actions = [
        'gerar_certificados', 
        'download_certificados_lote_action',
        'gerar_relatorio_excel',
        'gerar_relatorio_pdf'
    ]

    fieldsets = (
        ('Matrícula', {
            'fields': ('matricula',)
        }),
        ('Avaliação', {
            'fields': ('nota_final', 'frequencia', 'aprovado', 'observacoes')
        }),
        ('Certificado', {
            'fields': ('certificado_emitido', 'data_emissao_certificado')
        }),
        ('Auditoria', {
            'fields': ('avaliado_em', 'atualizado_em'),
            'classes': ('collapse',)
        }),
    )

    readonly_fields = ['avaliado_em', 'atualizado_em']

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related(
            'matricula__turma__evento',
            'matricula__interessado',
        )

    def get_numero_matricula(self, obj):
        return obj.matricula.numero_matricula
    get_numero_matricula.short_description = 'Matrícula'
    get_numero_matricula.admin_order_field = 'matricula__numero_matricula'

    def get_aluno(self, obj):
        return obj.matricula.interessado.nome
    get_aluno.short_description = 'Aluno'
    get_aluno.admin_order_field = 'matricula__interessado__nome'

    def get_turma(self, obj):
        return obj.matricula.turma
    get_turma.short_description = 'Turma'
    get_turma.admin_order_field = 'matricula__turma'

    def get_actions(self, request):
        actions = super().get_actions(request)
        if 'delete_selected' in actions:
            del actions['delete_selected']
        return actions

    def acoes_certificado(self, obj):
        if not obj.aprovado:
            return format_html('<span style="color: #999;">-</span>')
        preview_url = reverse('academico:preview_certificado', args=[obj.pk])
        return format_html(
            '<a href="{}" target="_blank" class="button" style="padding: 5px 12px; background-color: #417690; color: white; text-decoration: none; border-radius: 4px;">👁️ Ver</a>',
            preview_url
        )
    acoes_certificado.short_description = 'Certificado'

    def changelist_view(self, request, extra_context=None):
        from apps.eventos.models import Evento, Turma

        extra_context = extra_context or {}

        # Lista combinada de evento + turma que possuem avaliacoes
        turmas_com_avaliacao = Turma.objects.filter(
            matriculas__avaliacao__isnull=False
        ).select_related('evento').distinct().order_by('evento__nome', 'nome')

        extra_context['turmas_com_avaliacao'] = turmas_com_avaliacao

        # Estados dos filtros
        evento_filter = request.GET.get('matricula__turma__evento__id__exact', '')
        turma_filter = request.GET.get('matricula__turma__id__exact', '')
        aprovado_filter = request.GET.get('aprovado__exact', '')
        certificado_filter = request.GET.get('certificado_emitido__exact', '')

        extra_context['evento_selecionado'] = evento_filter
        extra_context['turma_selecionada'] = turma_filter
        extra_context['aprovado_selecionado'] = aprovado_filter
        extra_context['certificado_selecionado'] = certificado_filter

        # Nomes para o painel de contexto
        if turma_filter:
            try:
                turma = Turma.objects.select_related('evento').get(pk=turma_filter)
                extra_context['turma_nome'] = turma.nome
                extra_context['evento_nome'] = turma.evento.nome
            except Turma.DoesNotExist:
                pass

        return super().changelist_view(request, extra_context=extra_context)

    def _agrupar_por_turma(self, queryset):
        grupos = defaultdict(list)
        for avaliacao in queryset.select_related(
            'matricula__turma__evento', 'matricula__interessado'
        ).order_by(
            'matricula__turma__evento__nome',
            'matricula__turma__nome',
            'matricula__interessado__nome'
        ):
            turma = avaliacao.matricula.turma
            grupos[turma].append(avaliacao)
        return dict(grupos)

    # ==========================================
    # ACTION 1: GERAR RELATÓRIO EXCEL
    # ==========================================
    def gerar_relatorio_excel(self, request, queryset):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.drawing.image import Image as XLImage
        from django.conf import settings
        import os
        grupos_turma = self._agrupar_por_turma(queryset)
        if not grupos_turma:
            self.message_user(request, '❌ Nenhuma avaliação encontrada.', level=messages.ERROR)
            return
        wb = Workbook()
        wb.remove(wb.active)
        static_path = os.path.join(settings.BASE_DIR, 'static', 'images')
        brasao_path = os.path.join(static_path, 'brasao-2.png')
        logo_meta_path = os.path.join(static_path, 'favicon-metareciclagem.png')
        borda_fina = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for turma, avaliacoes in grupos_turma.items():
            sheet_name = f"{turma.evento.nome[:15]} - {turma.nome[:10]}"
            ws = wb.create_sheet(title=sheet_name)
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_margins.left = 0.7
            ws.page_margins.right = 0.7
            ws.page_margins.top = 0.75
            ws.page_margins.bottom = 0.75
            linha_atual = 1
            try:
                if os.path.exists(brasao_path):
                    img_b = XLImage(brasao_path)
                    img_b.width = 70
                    img_b.height = 70
                    ws.add_image(img_b, 'A1')
                if os.path.exists(logo_meta_path):
                    img_m = XLImage(logo_meta_path)
                    img_m.width = 70
                    img_m.height = 70
                    ws.add_image(img_m, 'G1')
            except Exception as e:
                print(f"Aviso: Imagens não carregadas - {e}")
            for col in ['A', 'G']:
                for row in range(1, 4):
                    ws[f'{col}{row}'].border = borda_fina
            ws.merge_cells('C1:E3')
            cell_meta = ws['C1']
            cell_meta.value = "METARECICLAGEM\n\nALUNOS MATRICULADOS"
            cell_meta.font = Font(name='Arial', size=14, bold=True, color="1F4788")
            cell_meta.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_meta.border = borda_fina
            ws.row_dimensions[1].height = 60
            linha_atual = 4
            ws.merge_cells(f'A{linha_atual}:H{linha_atual}')
            cell_info = ws[f'A{linha_atual}']
            cell_info.value = f"{turma.evento.nome}  |  {turma.nome}"
            cell_info.font = Font(name='Arial', size=11, bold=True)
            cell_info.alignment = Alignment(horizontal='center', vertical='center')
            cell_info.border = borda_fina
            ws.row_dimensions[linha_atual].height = 20
            linha_atual += 2
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 40
            fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            colunas = ['# Matrícula', 'Nome', 'Aprovado\n(S/N)', 'Nota', 'Frequência\n(%)', 'Observação']
            for col_idx, col_name in enumerate(colunas, start=1):
                cell = ws.cell(row=linha_atual, column=col_idx)
                cell.value = col_name
                cell.font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.fill = fill_header
                cell.border = borda_fina
            ws.row_dimensions[linha_atual].height = 35
            linha_atual += 1
            for avaliacao in avaliacoes:
                ws.cell(row=linha_atual, column=1).value = avaliacao.matricula.numero_matricula
                ws.cell(row=linha_atual, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=linha_atual, column=1).border = borda_fina
                ws.cell(row=linha_atual, column=2).value = avaliacao.matricula.interessado.nome
                ws.cell(row=linha_atual, column=2).alignment = Alignment(horizontal='left', vertical='center')
                ws.cell(row=linha_atual, column=2).border = borda_fina
                ws.cell(row=linha_atual, column=3).value = "Sim" if avaliacao.aprovado else ""
                ws.cell(row=linha_atual, column=3).font = Font(name='Arial', size=11, bold=True if avaliacao.aprovado else False)
                ws.cell(row=linha_atual, column=3).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=linha_atual, column=3).border = borda_fina
                ws.cell(row=linha_atual, column=4).value = avaliacao.nota_final if avaliacao.nota_final else ""
                ws.cell(row=linha_atual, column=4).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=linha_atual, column=4).border = borda_fina
                ws.cell(row=linha_atual, column=5).value = f"{avaliacao.frequencia}%" if avaliacao.frequencia else ""
                ws.cell(row=linha_atual, column=5).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=linha_atual, column=5).border = borda_fina
                ws.cell(row=linha_atual, column=6).value = ""
                ws.cell(row=linha_atual, column=6).alignment = Alignment(horizontal='left', vertical='center')
                ws.cell(row=linha_atual, column=6).border = borda_fina
                ws.row_dimensions[linha_atual].height = 20
                linha_atual += 1
            data_emissao = date.today().strftime("%d/%m/%Y")
            ws.oddFooter.left.text = f"Emitido em: {data_emissao}"
            ws.oddFooter.left.size = 9
            ws.oddFooter.right.text = "Página &P de &N"
            ws.oddFooter.right.size = 9
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Relatorio_Alunos_{date.today().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        self.message_user(
            request,
            f'✅ Relatório Excel gerado! {len(grupos_turma)} turma(s), {queryset.count()} aluno(s).',
            level=messages.SUCCESS
        )
        return response
    gerar_relatorio_excel.short_description = '📊 Gerar Relatório Excel (Segmentado por Turma)'

    # ==========================================
    # ACTION 2: GERAR RELATÓRIO PDF
    # ==========================================
    def gerar_relatorio_pdf(self, request, queryset):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from django.conf import settings
        import os
        grupos_turma = self._agrupar_por_turma(queryset)
        if not grupos_turma:
            self.message_user(request, '❌ Nenhuma avaliação encontrada.', level=messages.ERROR)
            return
        response = HttpResponse(content_type='application/pdf')
        filename = f"Relatorio_Alunos_{date.today().strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        doc = SimpleDocTemplate(
            response,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=1.5*cm,
            bottomMargin=2*cm,
        )
        story = []
        styles = getSampleStyleSheet()
        static_path = os.path.join(settings.BASE_DIR, 'static', 'images')
        primeira_turma = True
        for turma, avaliacoes in grupos_turma.items():
            if not primeira_turma:
                story.append(PageBreak())
            primeira_turma = False
            header_images = []
            try:
                brasao_path = os.path.join(static_path, 'brasao-2.png')
                if os.path.exists(brasao_path):
                    header_images.append(Image(brasao_path, width=2.2*cm, height=2.2*cm))
                else:
                    header_images.append("")
            except:
                header_images.append("")
            style_meta = ParagraphStyle(
                'MetaStyle',
                parent=styles['Heading1'],
                fontSize=13,
                textColor=colors.HexColor('#1F4788'),
                alignment=TA_CENTER,
                leading=16,
            )
            header_images.append(Paragraph("METARECICLAGEM<br/><br/><b>ALUNOS MATRICULADOS</b>", style_meta))
            try:
                logo_meta_path = os.path.join(static_path, 'favicon-metareciclagem.png')
                if os.path.exists(logo_meta_path):
                    header_images.append(Image(logo_meta_path, width=2.2*cm, height=2.2*cm))
                else:
                    header_images.append("")
            except:
                header_images.append("")
            if any(header_images):
                header_table = Table([header_images], colWidths=[3*cm, 11*cm, 3*cm])
                header_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                    ('ALIGN', (1, 0), (1, 0), 'CENTER'),
                    ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('BOX', (0, 0), (-1, -1), 1, colors.black),
                    ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ]))
                story.append(header_table)
                story.append(Spacer(1, 0.3*cm))
            style_info = ParagraphStyle(
                'InfoStyle',
                parent=styles['Normal'],
                fontSize=10,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )
            info_para = Paragraph(f"{turma.evento.nome}  |  {turma.nome}", style_info)
            info_table = Table([[info_para]], colWidths=[17*cm])
            info_table.setStyle(TableStyle([
                ('BOX', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            story.append(info_table)
            story.append(Spacer(1, 0.3*cm))
            data = [
                ['# Matrícula', 'Nome', 'Aprovado\n(S/N)', 'Nota', 'Freq.\n(%)', 'Observação']
            ]
            for avaliacao in avaliacoes:
                data.append([
                    avaliacao.matricula.numero_matricula,
                    avaliacao.matricula.interessado.nome,
                    "Sim" if avaliacao.aprovado else "",
                    str(avaliacao.nota_final) if avaliacao.nota_final else '',
                    f"{avaliacao.frequencia}%" if avaliacao.frequencia else '',
                    ''
                ])
            table = Table(data, colWidths=[2.5*cm, 6*cm, 1.5*cm, 1.5*cm, 1.5*cm, 4*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 1), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
                ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BOX', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ]))
            story.append(table)
        def add_page_number(canvas, doc):
            page_num = canvas.getPageNumber()
            canvas.saveState()
            canvas.setFont('Helvetica', 9)
            canvas.drawString(2*cm, 1.5*cm, f"Emitido em: {date.today().strftime('%d/%m/%Y')}")
            canvas.drawRightString(A4[0] - 2*cm, 1.5*cm, f"Página {page_num}")
            canvas.restoreState()
        doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
        self.message_user(
            request,
            f'✅ Relatório PDF gerado! {len(grupos_turma)} turma(s), {queryset.count()} aluno(s).',
            level=messages.SUCCESS
        )
        return response
    gerar_relatorio_pdf.short_description = '📄 Gerar Relatório PDF (Segmentado por Turma)'

    # ==========================================
    # ACTION 3: GERAR CERTIFICADOS
    # ==========================================
    def gerar_certificados(self, request, queryset):
        aprovados = queryset.filter(aprovado=True)
        if aprovados.count() == 0:
            self.message_user(request, '❌ Nenhum aluno aprovado foi selecionado.', level=messages.ERROR)
            return
        certificados_gerados = 0
        erros = []
        for avaliacao in aprovados:
            try:
                if not avaliacao.certificado_emitido:
                    avaliacao.certificado_emitido = True
                    avaliacao.data_emissao_certificado = date.today()
                    avaliacao.save()
                    certificados_gerados += 1
                else:
                    erros.append(f'{avaliacao.matricula.interessado.nome} já possui certificado emitido.')
            except Exception as e:
                erros.append(f'{avaliacao.matricula.interessado.nome}: {str(e)}')
        if certificados_gerados > 0:
            self.message_user(
                request,
                f'✅ {certificados_gerados} certificado(s) marcado(s) como emitido(s)!',
                level=messages.SUCCESS
            )
        if erros:
            for erro in erros:
                self.message_user(request, f'⚠️ {erro}', level=messages.WARNING)
    gerar_certificados.short_description = '✅ Marcar certificados como emitidos'

    # ==========================================
    # ACTION 4: DOWNLOAD CERTIFICADOS EM LOTE
    # ==========================================
    def download_certificados_lote_action(self, request, queryset):
        aprovados = queryset.filter(aprovado=True)
        if aprovados.count() == 0:
            self.message_user(request, '❌ Nenhum aluno aprovado foi selecionado.', level=messages.ERROR)
            return
        ids = ','.join(str(av.pk) for av in aprovados)
        url = reverse('academico:download_certificados_lote')
        return HttpResponseRedirect(f"{url}?ids={ids}")
    download_certificados_lote_action.short_description = '📦 Baixar certificados em lote (ZIP)'

    

    def aprovar_eventos(self, request, queryset):
        queryset.update(status='aprovado')

    def reprovar_eventos(self, request, queryset):
        queryset.update(status='reprovado')


