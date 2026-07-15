"""
Arquivo: reports.py
Caminho: apps/selecao/reports.py
Alteração: Adicionado parâmetro de ordenação (classificação ou nome) e atualizado logo
Data: 12/01/2026
"""

from django.http import HttpResponse
from django.conf import settings
from datetime import datetime
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.platypus.frames import Frame
from reportlab.platypus.doctemplate import PageTemplate, BaseDocTemplate
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from io import BytesIO
import os


class ReportTemplate(BaseDocTemplate):
    """Template customizado com cabeçalho e rodapé"""
    
    def __init__(self, *args, evento=None, tipo_relatorio='', ordem='classificacao', **kwargs):
        super().__init__(*args, **kwargs)
        self.evento = evento
        self.tipo_relatorio = tipo_relatorio
        self.ordem = ordem
        self.data_emissao = datetime.now()
        
    def build_header_footer(self, canvas, doc):
        """Constrói cabeçalho e rodapé em todas as páginas"""
        canvas.saveState()
        
        # Dimensões da página
        width, height = doc.pagesize
        
        # ============================================
        # CABEÇALHO COM FUNDO CINZA
        # ============================================
        
        # Fundo cinza do cabeçalho
        canvas.setFillColor(colors.HexColor('#e0e0e0'))
        canvas.rect(0, height - 2.5*cm, width, 2.5*cm, fill=True, stroke=False)
        
        # Brasão PMS (esquerda) - ATUALIZADO PARA brasao-horizontal.png
        try:
            brasao_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'brasao-horizontal.png')
            if os.path.exists(brasao_path):
                brasao = Image(brasao_path, width=3*cm, height=1.8*cm)
                brasao.drawOn(canvas, 0.5*cm, height - 2.3*cm)
        except Exception as e:
            print(f"Erro ao carregar brasão: {e}")
        
        # Logo Metareciclagem (direita)
        try:
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'favicon-metareciclagem.png')
            if os.path.exists(logo_path):
                logo = Image(logo_path, width=3*cm, height=1.8*cm)
                logo.drawOn(canvas, width - 3.5*cm, height - 2.3*cm)
        except Exception as e:
            print(f"Erro ao carregar logo: {e}")
        
        # Texto centralizado do cabeçalho
        canvas.setFont('Helvetica-Bold', 12)
        canvas.setFillColor(colors.HexColor('#205067'))
        
        # Título do relatório com indicação de ordenação
        if self.ordem == 'nome':
            texto_cabecalho = f"{self.tipo_relatorio} - Ordem: ALFABÉTICA"
        else:
            texto_cabecalho = f"{self.tipo_relatorio} - Ordem: CLASSIFICAÇÃO"
        
        canvas.drawCentredString(width/2, height - 1.2*cm, texto_cabecalho)
        
        # Nome do evento com data de classificação
        if self.evento:
            canvas.setFont('Helvetica', 10)
            
            # Buscar data da classificação (processado_em da primeira classificação)
            from apps.selecao.models import Classificacao
            primeira_class = Classificacao.objects.filter(
                inscricao__evento=self.evento
            ).order_by('processado_em').first()
            
            data_classificacao = ''
            if primeira_class and primeira_class.processado_em:
                data_classificacao = primeira_class.processado_em.strftime('%d/%m/%Y')
            
            if data_classificacao:
                texto_evento = f"{self.evento.nome} - Data da Classificação: {data_classificacao}"
            else:
                texto_evento = self.evento.nome
            
            canvas.drawCentredString(width/2, height - 1.8*cm, texto_evento)
        
        # ============================================
        # RODAPÉ
        # ============================================
        
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#666666'))
        
        # Data e hora de emissão (esquerda)
        data_hora = self.data_emissao.strftime('%d/%m/%Y às %H:%M')
        canvas.drawString(1*cm, 1*cm, f'Emissão: {data_hora}')
        
        # Número da página (direita)
        page_num = canvas.getPageNumber()
        total_pages = doc.page
        canvas.drawRightString(width - 1*cm, 1*cm, f'Página {page_num} de {total_pages}')
        
        canvas.restoreState()


class RelatorioAprovadosService:
    """
    Service para gerar relatórios PDF de aprovados usando ReportLab
    """
    
    @staticmethod
    def formatar_cpf(cpf):
        """Formata CPF com pontos e traço: 123.456.789-00"""
        if not cpf:
            return '—'
        # Remove caracteres não numéricos
        numeros = ''.join(filter(str.isdigit, cpf))
        if len(numeros) == 11:
            return f'{numeros[:3]}.{numeros[3:6]}.{numeros[6:9]}-{numeros[9:]}'
        return cpf
    
    @staticmethod
    def formatar_cpf_mascarado(cpf):
        """
        Formata CPF com máscara LGPD: 123.4**.***-**
        Mantém os 4 primeiros dígitos e mascara os últimos 7
        """
        if not cpf:
            return '—'
        # Remove caracteres não numéricos
        numeros = ''.join(filter(str.isdigit, cpf))
        if len(numeros) == 11:
            return f'{numeros[:3]}.{numeros[3]}**.***-**'
        return cpf
    
    @staticmethod
    def formatar_telefone(telefone):
        """Formata telefone: (99) 99999-0000 ou (99) 9999-0000"""
        if not telefone:
            return '—'
        # Remove caracteres não numéricos
        numeros = ''.join(filter(str.isdigit, telefone))
        
        if len(numeros) == 11:
            # Celular: (99) 99999-0000
            return f'({numeros[:2]}) {numeros[2:7]}-{numeros[7:]}'
        elif len(numeros) == 10:
            # Fixo: (99) 9999-0000
            return f'({numeros[:2]}) {numeros[2:6]}-{numeros[6:]}'
        else:
            return telefone
    
    @staticmethod
    def gerar_relatorio_staff(evento, classificacoes, ordem='classificacao'):
        """
        Gera relatório STAFF em PAISAGEM (com telefones e email)
        
        Args:
            evento: Objeto Evento
            classificacoes: QuerySet de Classificacao (já ordenado)
            ordem: 'classificacao' ou 'nome'
            
        Returns:
            HttpResponse com PDF
        """
        # Criar buffer
        buffer = BytesIO()
        
        # Criar documento em PAISAGEM
        doc = ReportTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=3*cm,  # Espaço para cabeçalho
            bottomMargin=2*cm,  # Espaço para rodapé
            evento=evento,
            tipo_relatorio='LISTA DE CONTATO - STAFF (CONFIDENCIAL)',
            ordem=ordem
        )
        
        # Definir template de página com cabeçalho/rodapé
        frame = Frame(
            doc.leftMargin, 
            doc.bottomMargin, 
            doc.width, 
            doc.height,
            id='normal'
        )
        template = PageTemplate(id='all', frames=frame, onPage=doc.build_header_footer)
        doc.addPageTemplates([template])
        
        # Conteúdo
        story = []
        
        # Espaçamento inicial
        story.append(Spacer(1, 0.5*cm))
        
        # Tabela de dados
        data = [['Pos.', 'St', 'Nome Completo', 'CPF', 'Celular', 'Telefone', 'Email', 'Pts', 'Observação']]
        
        for c in classificacoes:
            interessado = c.inscricao.interessado
            
            # Status: "A" para aprovado, "LE" para lista de espera
            if c.classificado:
                status = 'A'
            elif c.lista_espera:
                status = 'LE'
            else:
                status = '—'
            
            data.append([
                str(c.posicao),
                status,
                interessado.nome[:28],  # Limita tamanho
                RelatorioAprovadosService.formatar_cpf(interessado.cpf),
                RelatorioAprovadosService.formatar_telefone(interessado.celular),
                RelatorioAprovadosService.formatar_telefone(interessado.telefone),
                (interessado.email or '—')[:25],  # Limita tamanho
                f'{c.pontuacao_total:.1f}',
                ''  # Coluna observação vazia
            ])
        
        # Larguras das colunas (paisagem tem mais espaço)
        col_widths = [1*cm, 1*cm, 5*cm, 2.8*cm, 3*cm, 2.8*cm, 5*cm, 1.2*cm, 3*cm]
        
        table = Table(data, colWidths=col_widths)
        table.setStyle(TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#205067')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            
            # Corpo
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Posição
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),  # Status
            ('ALIGN', (3, 1), (3, -1), 'CENTER'),  # CPF
            ('ALIGN', (-2, 1), (-2, -1), 'CENTER'),  # Pontos
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 1), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
            
            # Linhas alternadas
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f9f9f9'), colors.white]),
            
            # Bordas
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#ccc')),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#154052')),
            ('INNERGRID', (0, 1), (-1, -1), 0.5, colors.HexColor('#ddd')),
        ]))
        
        story.append(table)
        
        # Build PDF
        doc.build(story)
        
        # Retornar resposta
        pdf = buffer.getvalue()
        buffer.close()
        
        tipo_ordem = 'nome' if ordem == 'nome' else 'classificacao'
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="staff_{tipo_ordem}_{evento.nome.replace(" ", "_")}.pdf"'
        
        return response
    
    @staticmethod
    def gerar_relatorio_mural(evento, classificacoes, ordem='classificacao'):
        """
        Gera relatório MURAL em RETRATO (público, sem telefones, CPF mascarado)
        
        Args:
            evento: Objeto Evento
            classificacoes: QuerySet de Classificacao (já ordenado)
            ordem: 'classificacao' ou 'nome'
            
        Returns:
            HttpResponse com PDF
        """
        # Criar buffer
        buffer = BytesIO()
        
        # Criar documento em RETRATO
        doc = ReportTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=3.5*cm,  # Espaço para cabeçalho
            bottomMargin=2*cm,  # Espaço para rodapé
            evento=evento,
            tipo_relatorio='LISTA DE CLASSIFICADOS',
            ordem=ordem
        )
        
        # Definir template de página com cabeçalho/rodapé
        frame = Frame(
            doc.leftMargin, 
            doc.bottomMargin, 
            doc.width, 
            doc.height,
            id='normal'
        )
        template = PageTemplate(id='all', frames=frame, onPage=doc.build_header_footer)
        doc.addPageTemplates([template])
        
        # Conteúdo
        story = []
        
        # Espaçamento inicial
        story.append(Spacer(1, 0.8*cm))
        
        # Tabela de dados
        data = [['Posição', 'Status', 'Nome Completo', 'CPF', 'Pontuação']]
        
        for c in classificacoes:
            interessado = c.inscricao.interessado
            
            # Status: "Aprovado" ou "Espera"
            if c.classificado:
                status = 'Aprovado'
            elif c.lista_espera:
                status = 'Espera'
            else:
                status = '—'
            
            data.append([
                f'{c.posicao}º',
                status,
                interessado.nome,
                RelatorioAprovadosService.formatar_cpf_mascarado(interessado.cpf),
                f'{c.pontuacao_total:.2f}'
            ])
        
        col_widths = [2*cm, 2.5*cm, 8*cm, 3*cm, 2.5*cm]
        
        table = Table(data, colWidths=col_widths)
        table.setStyle(TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#205067')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Corpo
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Posição
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),  # Status
            ('ALIGN', (3, 1), (3, -1), 'CENTER'),  # CPF
            ('ALIGN', (4, 1), (4, -1), 'CENTER'),  # Pontuação
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
            
            # Linhas alternadas
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f9f9f9'), colors.white]),
            
            # Bordas
            ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#ccc')),
            ('LINEBELOW', (0, 0), (-1, 0), 3, colors.HexColor('#154052')),
            ('INNERGRID', (0, 1), (-1, -1), 1, colors.HexColor('#ddd')),
        ]))
        
        story.append(table)
        story.append(Spacer(1, 1*cm))
        
        # Aviso
        styles = getSampleStyleSheet()
        aviso_style = ParagraphStyle(
            'Aviso',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#004085'),
            spaceAfter=8,
            alignment=TA_CENTER,
            fontName='Helvetica'
        )
        
        story.append(Paragraph('<b>📞 ATENÇÃO CLASSIFICADOS!</b>', aviso_style))
        story.append(Paragraph('Os aprovados receberão contato da equipe nos próximos dias.', aviso_style))
        story.append(Paragraph('Mantenha seus telefones atualizados e aguarde nossa ligação.', aviso_style))
        
        # Build PDF
        doc.build(story)
        
        # Retornar resposta
        pdf = buffer.getvalue()
        buffer.close()
        
        tipo_ordem = 'nome' if ordem == 'nome' else 'classificacao'
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="mural_{tipo_ordem}_{evento.nome.replace(" ", "_")}.pdf"'
        
        return response
    
        # ======================
    # EXPORTAÇÃO PARA EXCEL
    # Adicionado em 12/01/2026
    # ======================
    
    @staticmethod
    def gerar_excel_staff(evento, classificacoes, ordem='classificacao'):
        """
        Gera planilha Excel STAFF (com telefones e email)
        
        Args:
            evento: Objeto Evento
            classificacoes: QuerySet de Classificacao (já ordenado)
            ordem: 'classificacao' ou 'nome'
            
        Returns:
            HttpResponse com Excel
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Lista Contato Staff"
        
        # Buscar data da classificação
        from apps.selecao.models import Classificacao as ClassModel
        primeira_class = ClassModel.objects.filter(
            inscricao__evento=evento
        ).order_by('processado_em').first()
        
        data_classificacao = ''
        if primeira_class and primeira_class.processado_em:
            data_classificacao = primeira_class.processado_em.strftime('%d/%m/%Y')
        
        # Título
        tipo_ordem = 'ALFABÉTICA' if ordem == 'nome' else 'CLASSIFICAÇÃO'
        ws.merge_cells('A1:I1')
        titulo_cell = ws['A1']
        titulo_cell.value = f'LISTA DE CONTATO - STAFF (CONFIDENCIAL) - Ordem: {tipo_ordem}'
        titulo_cell.font = Font(size=14, bold=True, color='FFFFFF')
        titulo_cell.fill = PatternFill(start_color='205067', end_color='205067', fill_type='solid')
        titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Subtítulo com evento e data
        ws.merge_cells('A2:I2')
        subtitulo_cell = ws['A2']
        if data_classificacao:
            subtitulo_cell.value = f'{evento.nome} - Data da Classificação: {data_classificacao}'
        else:
            subtitulo_cell.value = evento.nome
        subtitulo_cell.font = Font(size=12, bold=True)
        subtitulo_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 25
        
        # Linha em branco
        ws.row_dimensions[3].height = 10
        
        # Cabeçalhos
        headers = ['Posição', 'Status', 'Nome Completo', 'CPF', 'Celular', 'Telefone', 'Email', 'Pontos', 'Observação']
        header_row = 4
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='205067', end_color='205067', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        ws.row_dimensions[header_row].height = 20
        
        # Dados
        row_num = header_row + 1
        for c in classificacoes:
            interessado = c.inscricao.interessado
            
            # Status
            if c.classificado:
                status = 'A'
            elif c.lista_espera:
                status = 'LE'
            else:
                status = '—'
            
            # Dados da linha
            data_row = [
                c.posicao,
                status,
                interessado.nome,
                RelatorioAprovadosService.formatar_cpf(interessado.cpf),
                RelatorioAprovadosService.formatar_telefone(interessado.celular),
                RelatorioAprovadosService.formatar_telefone(interessado.telefone),
                interessado.email or '—',
                float(c.pontuacao_total),
                ''  # Observação vazia
            ]
            
            for col_num, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal='center' if col_num in [1, 2, 4, 8] else 'left', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Cor de fundo alternada
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color='F9F9F9', end_color='F9F9F9', fill_type='solid')
            
            ws.row_dimensions[row_num].height = 18
            row_num += 1
        
        # Ajustar largura das colunas
        column_widths = [10, 8, 35, 18, 18, 18, 35, 10, 25]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Salvar em buffer
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Retornar resposta
        tipo_ordem_file = 'nome' if ordem == 'nome' else 'classificacao'
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="staff_{tipo_ordem_file}_{evento.nome.replace(" ", "_")}.xlsx"'
        
        return response
    
    @staticmethod
    def gerar_excel_mural(evento, classificacoes, ordem='classificacao'):
        """
        Gera planilha Excel MURAL (público, sem telefones, CPF mascarado)
        
        Args:
            evento: Objeto Evento
            classificacoes: QuerySet de Classificacao (já ordenado)
            ordem: 'classificacao' ou 'nome'
            
        Returns:
            HttpResponse com Excel
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Lista Classificados"
        
        # Buscar data da classificação
        from apps.selecao.models import Classificacao as ClassModel
        primeira_class = ClassModel.objects.filter(
            inscricao__evento=evento
        ).order_by('processado_em').first()
        
        data_classificacao = ''
        if primeira_class and primeira_class.processado_em:
            data_classificacao = primeira_class.processado_em.strftime('%d/%m/%Y')
        
        # Título
        tipo_ordem = 'ALFABÉTICA' if ordem == 'nome' else 'CLASSIFICAÇÃO'
        ws.merge_cells('A1:E1')
        titulo_cell = ws['A1']
        titulo_cell.value = f'LISTA DE CLASSIFICADOS - Ordem: {tipo_ordem}'
        titulo_cell.font = Font(size=14, bold=True, color='FFFFFF')
        titulo_cell.fill = PatternFill(start_color='205067', end_color='205067', fill_type='solid')
        titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Subtítulo com evento e data
        ws.merge_cells('A2:E2')
        subtitulo_cell = ws['A2']
        if data_classificacao:
            subtitulo_cell.value = f'{evento.nome} - Data da Classificação: {data_classificacao}'
        else:
            subtitulo_cell.value = evento.nome
        subtitulo_cell.font = Font(size=12, bold=True)
        subtitulo_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 25
        
        # Linha em branco
        ws.row_dimensions[3].height = 10
        
        # Cabeçalhos
        headers = ['Posição', 'Status', 'Nome Completo', 'CPF', 'Pontuação']
        header_row = 4
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='205067', end_color='205067', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        ws.row_dimensions[header_row].height = 20
        
        # Dados
        row_num = header_row + 1
        for c in classificacoes:
            interessado = c.inscricao.interessado
            
            # Status
            if c.classificado:
                status = 'Aprovado'
            elif c.lista_espera:
                status = 'Espera'
            else:
                status = '—'
            
            # Dados da linha
            data_row = [
                f'{c.posicao}º',
                status,
                interessado.nome,
                RelatorioAprovadosService.formatar_cpf_mascarado(interessado.cpf),
                float(c.pontuacao_total)
            ]
            
            for col_num, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal='center' if col_num in [1, 2, 4, 5] else 'left', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Cor de fundo alternada
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color='F9F9F9', end_color='F9F9F9', fill_type='solid')
            
            ws.row_dimensions[row_num].height = 18
            row_num += 1
        
        # Ajustar largura das colunas
        column_widths = [12, 15, 40, 20, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Adicionar aviso no final
        aviso_row = row_num + 2
        ws.merge_cells(f'A{aviso_row}:E{aviso_row}')
        aviso_cell = ws[f'A{aviso_row}']
        aviso_cell.value = '📞 ATENÇÃO CLASSIFICADOS! Os aprovados receberão contato da equipe nos próximos dias.'
        aviso_cell.font = Font(size=11, bold=True, color='004085')
        aviso_cell.fill = PatternFill(start_color='D4EDFF', end_color='D4EDFF', fill_type='solid')
        aviso_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[aviso_row].height = 40
        
        # Salvar em buffer
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Retornar resposta
        tipo_ordem_file = 'nome' if ordem == 'nome' else 'classificacao'
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="mural_{tipo_ordem_file}_{evento.nome.replace(" ", "_")}.xlsx"'
        
        return response
    
